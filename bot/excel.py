import pandas as pd
import os
import time
import difflib
from collections import defaultdict
from openpyxl import load_workbook

try:
    from lector_excel import cargar_pedido, hojas
    from descargar_dropbox import descargar_excels, actualizar_archivo
    rel = ""
except Exception:
    rel = "bot/"
    from bot.lector_excel import cargar_pedido, hojas
    from bot.descargar_dropbox import descaergar_excels, actualizar_archivo

def excel_to_table(file, hoja):
    table = pd.read_excel(
        file, sheetname=hoja, header=0,
        parse_cols="A, B, C, D, E, F, G, H, I, J, K, L, M, N, O, P, Q," + \
        " R, S, T, U, V, W, X, Y, Z")
    fin = 0
    for i in range(len(table)):
        fila = table.iloc[i].real
        nombre = fila[0]
        fin = i
        if str(nombre) == " " or str(nombre) == "nan" or str(nombre) == "":
            break
    tabla = table[0:fin]
    tabla = [[j for j in tabla.iloc[i].real]
             for i in range(len(tabla))]
    return tabla

def cargar_pedidos(path):
    pedidos_totales = {}
    files = [f for f in os.listdir(path) if f[0:2] == "20" and ".xl" in f]
    files.remove("2017 - 08 (Agosto).xlsx")
    for file in files:
        sheets = hojas(path + "/" + file)[1:]
        for hoja in sheets:
            pedido, info = cargar_pedido(path + "/" + file, hoja)
            pedidos_totales[info["fecha_formato"]] = pedido
    return pedidos_totales

def deudas_jefes(jefes, path=rel+"datos"):
    pedidos = cargar_pedidos(path)
    deudas = defaultdict(list)
    for fecha in sorted(list(pedidos.keys())):
        for persona in pedidos[fecha]:
            if persona["nombre"] in jefes:
                if persona["pagado"] == "NO":
                    deudas[persona["nombre"]].append(
                        (fecha.replace("-", "/"), persona["deuda"]))
    return deudas

def parser_to_excel(parser, mes, pedido):
    descargar_excels("datos")
    file = [f for f in os.listdir(rel+"datos")
             if f[0:2] == "20" and ".xl" in f and mes in f][0]
    wb = load_workbook(filename=rel+"datos/"+file, keep_vba=True)
    ws = wb[pedido]
    columnas = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ") + ["AA", "AB", "AC", "AD"]
    productos = []
    productos_index = []
    agregar = False
    for col in columnas:
        valor = ws["{}4".format(col)].value
        if valor == "Deuda":
            agregar = True
        if valor == "Total":
            total_index = col
            break
        if agregar:
            productos.append(valor)
            productos_index.append(col)
    dic_index = dict(zip(productos, productos_index))
    prod_parser = {}
    for p in parser.productos:
        try:
            match = difflib.get_close_matches(p, productos)[0]
            prod_parser[p] = match
        except IndexError:
            raise Exception("No se ha encontrado el producto {}".format(p))
    i = 4
    encontrados = []
    while True:
        nombre, apellido = ws["B{}".format(i)].value, ws["C{}".format(i)].value
        nombre_completo = nombre + " " + apellido
        for jefe in parser.jefes:
            if jefe == nombre_completo:
                pedido = parser.jefes[jefe]
                for p, c in pedido.items():
                    col = dic_index[prod_parser[p]]
                    ws["{}{}".format(col, i)] = c
                encontrados.append(jefe)
        i += 1
        if ws["B{}".format(i)].value is None:
            break
    no_encontrados = [j for j in parser.jefes if j not in encontrados]
    if len(no_encontrados) > 0:
        raise Exception("{} no encontrado".format(jefe))
    path = rel + "datos/" + file
    wb.save(path)
    actualizar_archivo(path, "/Frutillas/Contabilidad/2017 - 2")
    
            

if __name__ == "__main__":
    d = deudas_jefes(["Majo Castañeda"])
    print(d)
